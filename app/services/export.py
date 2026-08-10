import io
from datetime import datetime
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from app.models.registration import get_registrations, count_registrations_for_event


def export_excel(event):
    """Generate and return an Excel file with registrations for the given event."""
    # event is a dict from get_campus_event()
    wb = Workbook()
    ws = wb.active
    ws.title = '报名表'

    # Header styles
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    registrations = get_registrations(event['id'])
    reg_count = count_registrations_for_event(event['id'])

    # Title row
    ws.merge_cells('A1:H1')
    ws['A1'] = f'「{event["title"]}」报名表'
    ws['A1'].font = Font(name='微软雅黑', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Sub-info row
    ws.merge_cells('A2:H2')
    ws['A2'] = f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}  |  报名人数: {reg_count}人'
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['序号', '用户名', '姓名', '年级', '联系电话', '联系邮箱', '备注', '报名时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for i, reg in enumerate(registrations, 1):
        reg_time = reg['registered_at']
        if isinstance(reg_time, str):
            reg_time_str = reg_time[:16] if len(reg_time) >= 16 else reg_time
        else:
            reg_time_str = reg_time.strftime('%Y-%m-%d %H:%M')

        row_data = [
            i,
            reg.get('username', ''),
            reg.get('real_name', '') or '',
            reg.get('grade', '') or '',
            reg.get('contact_phone', '') or '',
            reg.get('contact_email', '') or '',
            reg.get('notes', '') or '',
            reg_time_str,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=4 + i, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # Column widths
    widths = [6, 15, 12, 10, 16, 22, 25, 18]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'{event["title"]}_报名表_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
